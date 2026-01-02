#!/usr/bin/env python3
"""GTD Excel Generator - Clean and Strategic"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from gtd_data import gtd_items

# Create DataFrame
df = pd.DataFrame(gtd_items)
df.columns = ['Item', 'Category', 'Project/Area', 'Context', 'Next Action', 'Waiting For', 'Someday/Maybe', 'Priority', 'Status', 'Energy', 'Time Required', 'Due Date', 'Delegated To', 'Notes']

# Create workbook
wb = Workbook()

# Styles
COLORS = {
    'header': '2C3E50', 'critical': 'FF6B6B', 'high': 'FFA94D', 'medium': 'FFD93D', 'low': '6BCB77',
    'next_action': 'FF6B6B', 'project': '4ECDC4', 'waiting': 'A8E6CF', 'someday': 'DDA0DD',
    'recurring_daily': 'FFB6C1', 'recurring_weekly': 'E6E6FA', 'recurring_monthly': 'B0E0E6',
    'recurring_quarterly': 'F0E68C', 'ongoing': '87CEEB', 'professional': '4A90A4', 'personal': '9B59B6'
}
header_font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
body_font = Font(name='Aptos', size=10)
thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                     top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

def create_sheet(ws, data, header_color='2C3E50'):
    """Create a formatted sheet with data"""
    headers = list(data.columns)
    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    # Data rows
    for row_num, row_data in enumerate(data.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value if pd.notna(value) else '')
            cell.font = body_font
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border
            # Color coding
            col_name = headers[col_num-1]
            if col_name == 'Priority':
                colors = {'Critical': ('FF6B6B', 'FFFFFF'), 'High': ('FFA94D', '000000'), 'Medium': ('FFD93D', '000000'), 'Low': ('6BCB77', '000000')}
                if value in colors:
                    cell.fill = PatternFill(start_color=colors[value][0], end_color=colors[value][0], fill_type='solid')
                    cell.font = Font(name='Aptos', size=10, bold=True, color=colors[value][1])
            elif col_name == 'Status':
                status_colors = {'Next Action': 'FF6B6B', 'Project': '4ECDC4', 'Waiting For': 'A8E6CF', 'Someday/Maybe': 'DDA0DD',
                                'Recurring - Daily': 'FFB6C1', 'Recurring - Weekly': 'E6E6FA', 'Recurring - Monthly': 'B0E0E6',
                                'Recurring - Quarterly': 'F0E68C', 'Ongoing': '87CEEB', 'As Needed': 'F5DEB3', 'Reference': 'D3D3D3'}
                if value in status_colors:
                    cell.fill = PatternFill(start_color=status_colors[value], end_color=status_colors[value], fill_type='solid')
            elif col_name == 'Category':
                if value == 'Professional':
                    cell.fill = PatternFill(start_color='4A90A4', end_color='4A90A4', fill_type='solid')
                    cell.font = Font(name='Aptos', size=10, color='FFFFFF')
                elif value == 'Personal':
                    cell.fill = PatternFill(start_color='9B59B6', end_color='9B59B6', fill_type='solid')
                    cell.font = Font(name='Aptos', size=10, color='FFFFFF')
    # Column widths
    widths = {'Item': 50, 'Category': 12, 'Project/Area': 28, 'Context': 12, 'Next Action': 45, 'Waiting For': 25,
              'Someday/Maybe': 5, 'Priority': 10, 'Status': 18, 'Energy': 10, 'Time Required': 12, 'Due Date': 10, 'Delegated To': 12, 'Notes': 40}
    for col, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(header, 15)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"

# 1. GTD Master List
ws = wb.active
ws.title = "GTD Master List"
create_sheet(ws, df)

# 2. Next Actions
ws_next = wb.create_sheet("Next Actions")
create_sheet(ws_next, df[df['Status'] == 'Next Action'], 'FF6B6B')

# 3. Projects
ws_proj = wb.create_sheet("Projects")
create_sheet(ws_proj, df[df['Status'] == 'Project'], '4ECDC4')

# 4. Waiting For
ws_wait = wb.create_sheet("Waiting For")
create_sheet(ws_wait, df[df['Status'] == 'Waiting For'], 'A8E6CF')

# 5. Someday/Maybe
ws_someday = wb.create_sheet("Someday-Maybe")
create_sheet(ws_someday, df[df['Status'] == 'Someday/Maybe'], 'DDA0DD')

# 6. Recurring
ws_recurring = wb.create_sheet("Recurring Tasks")
create_sheet(ws_recurring, df[df['Status'].str.contains('Recurring', na=False)], 'E6E6FA')

# 7. High Priority
ws_priority = wb.create_sheet("High Priority")
create_sheet(ws_priority, df[df['Priority'].isin(['Critical', 'High'])], 'FF6B6B')

# 8. Professional
ws_prof = wb.create_sheet("Professional")
create_sheet(ws_prof, df[df['Category'] == 'Professional'], '4A90A4')

# 9. Personal
ws_pers = wb.create_sheet("Personal")
create_sheet(ws_pers, df[df['Category'] == 'Personal'], '9B59B6')

# 10. AI Skills & Development (NEW - Core Task)
ws_ai = wb.create_sheet("AI Skills - CORE")
ai_items = df[df['Project/Area'].str.contains('AI', case=False, na=False)]
create_sheet(ws_ai, ai_items, 'FF6B6B')

# 11. Strategic Accounts
ws_accounts = wb.create_sheet("Strategic Accounts")
accounts = ['World Bank', 'HPE', 'CAA', 'BAT', 'Riyadh Air', 'AstraZeneca']
account_items = df[df['Project/Area'].str.contains('|'.join(accounts), case=False, na=False, regex=True)]
create_sheet(ws_accounts, account_items, '4A90A4')

# 12. Business Planning
ws_biz = wb.create_sheet("Business Planning")
biz_items = df[df['Project/Area'].str.contains('Business|Professional Identity|Personal Projects', case=False, na=False, regex=True)]
create_sheet(ws_biz, biz_items, '2C3E50')

# 13. Family & Health
ws_family = wb.create_sheet("Family & Health")
family_items = df[df['Project/Area'].str.contains('Family|Health|Mindfulness', case=False, na=False, regex=True)]
create_sheet(ws_family, family_items, '9B59B6')

# 14. Weekly Review
ws_review = wb.create_sheet("Weekly Review")
cell = ws_review.cell(row=1, column=1, value="WEEKLY REVIEW - Shreyas GTD")
cell.font = Font(name='Aptos', size=16, bold=True)
ws_review.merge_cells('A1:D1')
row = 3
sections = [
    ("🔴 CRITICAL - Address This Week", df[df['Priority'] == 'Critical']),
    ("🟠 HIGH PRIORITY", df[df['Priority'] == 'High'].head(15)),
    ("⏳ WAITING FOR - Follow Up", df[df['Status'] == 'Waiting For']),
    ("🤖 AI SKILLS - Daily Practice", df[df['Project/Area'].str.contains('AI Skills', case=False, na=False)]),
    ("🔄 RECURRING - Weekly", df[df['Status'] == 'Recurring - Weekly']),
]
for title, items in sections:
    if len(items) == 0: continue
    ws_review.cell(row=row, column=1, value=title).font = Font(name='Aptos', size=12, bold=True)
    row += 1
    for _, item in items.iterrows():
        ws_review.cell(row=row, column=1, value=f"□ {item['Item']}").font = body_font
        ws_review.cell(row=row, column=2, value=item['Project/Area']).font = body_font
        ws_review.cell(row=row, column=3, value=item['Next Action']).font = body_font
        row += 1
    row += 1
ws_review.column_dimensions['A'].width = 50
ws_review.column_dimensions['B'].width = 25
ws_review.column_dimensions['C'].width = 45

# 15. Legend
ws_legend = wb.create_sheet("Legend")
legends = [
    ("Priority Colors:", [("Critical", "FF6B6B"), ("High", "FFA94D"), ("Medium", "FFD93D"), ("Low", "6BCB77")]),
    ("Status Colors:", [("Next Action", "FF6B6B"), ("Project", "4ECDC4"), ("Waiting For", "A8E6CF"), ("Someday/Maybe", "DDA0DD"), ("Recurring", "E6E6FA")]),
    ("Contexts:", [("@Computer", ""), ("@Phone", ""), ("@Office", ""), ("@Home", ""), ("@Thinking", ""), ("@Errands", ""), ("@Anywhere", "")])
]
row = 1
for section, items in legends:
    ws_legend.cell(row=row, column=1, value=section).font = Font(name='Aptos', size=12, bold=True)
    row += 1
    for name, color in items:
        cell = ws_legend.cell(row=row, column=1, value=name)
        cell.font = body_font
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        row += 1
    row += 1
ws_legend.column_dimensions['A'].width = 25

# Save
output_path = '/Users/shreyas/Downloads/GTD-Gather/Shreyas_GTD_Master.xlsx'
wb.save(output_path)

print(f"✅ GTD Excel created: {output_path}")
print(f"\n📊 Summary:")
print(f"   Total items: {len(df)}")
print(f"   Professional: {len(df[df['Category']=='Professional'])}")
print(f"   Personal: {len(df[df['Category']=='Personal'])}")
print(f"   Critical: {len(df[df['Priority']=='Critical'])}")
print(f"   High: {len(df[df['Priority']=='High'])}")
print(f"   AI Skills items: {len(ai_items)}")
print(f"\n📋 15 Sheets created - Ready for GTD processing!")
