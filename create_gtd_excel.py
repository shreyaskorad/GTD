#!/usr/bin/env python3
"""
GTD Master List Excel Generator for Shreyas
Creates a comprehensive GTD-formatted Excel with color coding, filters, and formatting
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# Read the CSV data
df = pd.read_csv('/Users/shreyas/Downloads/GTD-Gather/Shreyas_GTD_Master_List.csv')

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "GTD Master List"

# Define colors
colors = {
    # Priority colors
    'Critical': 'FF6B6B',      # Red
    'High': 'FFA94D',          # Orange
    'Medium': 'FFD93D',        # Yellow
    'Low': '6BCB77',           # Green
    
    # Status colors
    'Next Action': 'FF6B6B',   # Red - Do it
    'Project': '4ECDC4',       # Teal - Multi-step
    'Waiting For': 'A8E6CF',   # Light green - Delegated
    'Someday/Maybe': 'DDA0DD', # Plum - Future
    'Recurring - Daily': 'FFB6C1',   # Light pink
    'Recurring - Weekly': 'E6E6FA',  # Lavender
    'Recurring - Monthly': 'B0E0E6', # Powder blue
    'Recurring - Quarterly': 'F0E68C', # Khaki
    'Recurring - Yearly': 'DEB887',  # Burlywood
    'Ongoing': '87CEEB',       # Sky blue
    'As Needed': 'F5DEB3',     # Wheat
    'Reference': 'D3D3D3',     # Light gray
    
    # Category colors
    'Professional': '4A90A4',  # Steel blue
    'Personal': '9B59B6',      # Purple
    
    # Header
    'Header': '2C3E50',        # Dark blue-gray
}

# Define Aptos font (fallback to Calibri if Aptos not available)
header_font = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
body_font = Font(name='Aptos', size=10)
bold_font = Font(name='Aptos', size=10, bold=True)

# Define borders
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Write header
headers = list(df.columns)
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color=colors['Header'], end_color=colors['Header'], fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Write data rows
for row_num, row_data in enumerate(df.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value if pd.notna(value) else '')
        cell.font = body_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border
        
        # Apply category coloring (column 2)
        if col_num == 2:  # Category column
            if value == 'Professional':
                cell.fill = PatternFill(start_color=colors['Professional'], end_color=colors['Professional'], fill_type='solid')
                cell.font = Font(name='Aptos', size=10, color='FFFFFF')
            elif value == 'Personal':
                cell.fill = PatternFill(start_color=colors['Personal'], end_color=colors['Personal'], fill_type='solid')
                cell.font = Font(name='Aptos', size=10, color='FFFFFF')
        
        # Apply priority coloring (column 8)
        if col_num == 8:  # Priority column
            if value == 'Critical':
                cell.fill = PatternFill(start_color=colors['Critical'], end_color=colors['Critical'], fill_type='solid')
                cell.font = Font(name='Aptos', size=10, bold=True, color='FFFFFF')
            elif value == 'High':
                cell.fill = PatternFill(start_color=colors['High'], end_color=colors['High'], fill_type='solid')
                cell.font = Font(name='Aptos', size=10, bold=True)
            elif value == 'Medium':
                cell.fill = PatternFill(start_color=colors['Medium'], end_color=colors['Medium'], fill_type='solid')
            elif value == 'Low':
                cell.fill = PatternFill(start_color=colors['Low'], end_color=colors['Low'], fill_type='solid')
        
        # Apply status coloring (column 9)
        if col_num == 9:  # Status column
            status_colors = {
                'Next Action': ('FF6B6B', 'FFFFFF'),
                'Project': ('4ECDC4', 'FFFFFF'),
                'Waiting For': ('A8E6CF', '000000'),
                'Someday/Maybe': ('DDA0DD', '000000'),
                'Recurring - Daily': ('FFB6C1', '000000'),
                'Recurring - Weekly': ('E6E6FA', '000000'),
                'Recurring - Monthly': ('B0E0E6', '000000'),
                'Recurring - Quarterly': ('F0E68C', '000000'),
                'Recurring - Yearly': ('DEB887', '000000'),
                'Ongoing': ('87CEEB', '000000'),
                'As Needed': ('F5DEB3', '000000'),
                'Reference': ('D3D3D3', '000000'),
            }
            if value in status_colors:
                bg_color, font_color = status_colors[value]
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.font = Font(name='Aptos', size=10, color=font_color)

# Set column widths
column_widths = {
    'A': 55,  # Item
    'B': 12,  # Category
    'C': 30,  # Project/Area
    'D': 12,  # Context
    'E': 50,  # Next Action
    'F': 35,  # Waiting For
    'G': 5,   # Someday/Maybe
    'H': 10,  # Priority
    'I': 18,  # Status
    'J': 10,  # Energy Level
    'K': 12,  # Time Required
    'L': 12,  # Due Date
    'M': 15,  # Delegated To
    'N': 50,  # Notes
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze top row
ws.freeze_panes = 'A2'

# Add auto-filter
ws.auto_filter.ref = f"A1:N{len(df) + 1}"

# Set row height for header
ws.row_dimensions[1].height = 25

# ============================================
# CREATE ADDITIONAL SHEETS
# ============================================

# --- NEXT ACTIONS SHEET ---
ws_next = wb.create_sheet("Next Actions")
next_actions = df[df['Status'] == 'Next Action']

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws_next.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Write data
for row_num, row_data in enumerate(next_actions.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_next.cell(row=row_num, column=col_num, value=value if pd.notna(value) else '')
        cell.font = body_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border

for col_letter, width in column_widths.items():
    ws_next.column_dimensions[col_letter].width = width
ws_next.freeze_panes = 'A2'
ws_next.auto_filter.ref = f"A1:N{len(next_actions) + 1}"

# --- PROJECTS SHEET ---
ws_proj = wb.create_sheet("Projects")
projects = df[df['Status'] == 'Project']

for col_num, header in enumerate(headers, 1):
    cell = ws_proj.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color='4ECDC4', end_color='4ECDC4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for row_num, row_data in enumerate(projects.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_proj.cell(row=row_num, column=col_num, value=value if pd.notna(value) else '')
        cell.font = body_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border

for col_letter, width in column_widths.items():
    ws_proj.column_dimensions[col_letter].width = width
ws_proj.freeze_panes = 'A2'
ws_proj.auto_filter.ref = f"A1:N{len(projects) + 1}"

# --- WAITING FOR SHEET ---
ws_wait = wb.create_sheet("Waiting For")
waiting = df[df['Status'] == 'Waiting For']

for col_num, header in enumerate(headers, 1):
    cell = ws_wait.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color='A8E6CF', end_color='A8E6CF', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for row_num, row_data in enumerate(waiting.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_wait.cell(row=row_num, column=col_num, value=value if pd.notna(value) else '')
        cell.font = body_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border

for col_letter, width in column_widths.items():
    ws_wait.column_dimensions[col_letter].width = width
ws_wait.freeze_panes = 'A2'
ws_wait.auto_filter.ref = f"A1:N{len(waiting) + 1}"

# --- SOMEDAY/MAYBE SHEET ---
ws_someday = wb.create_sheet("Someday-Maybe")
someday = df[df['Status'] == 'Someday/Maybe']

for col_num, header in enumerate(headers, 1):
    cell = ws_someday.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color='DDA0DD', end_color='DDA0DD', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for row_num, row_data in enumerate(someday.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_someday.cell(row=row_num, column=col_num, value=value if pd.notna(value) else '')
        cell.font = body_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border

for col_letter, width in column_widths.items():
    ws_someday.column_dimensions[col_letter].width = width
ws_someday.freeze_panes = 'A2'
ws_someday.auto_filter.ref = f"A1:N{len(someday) + 1}"

# --- RECURRING TASKS SHEET ---
ws_recurring = wb.create_sheet("Recurring Tasks")
recurring = df[df['Status'].str.contains('Recurring', na=False)]

for col_num, header in enumerate(headers, 1):
    cell = ws_recurring.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for row_num, row_data in enumerate(recurring.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_recurring.cell(row=row_num, column=col_num, value=value if pd.notna(value) else '')
        cell.font = body_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border
        # Color code by frequency
        if col_num == 9:  # Status column
            if 'Daily' in str(value):
                cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            elif 'Weekly' in str(value):
                cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            elif 'Monthly' in str(value):
                cell.fill = PatternFill(start_color='B0E0E6', end_color='B0E0E6', fill_type='solid')
            elif 'Quarterly' in str(value):
                cell.fill = PatternFill(start_color='F0E68C', end_color='F0E68C', fill_type='solid')
            elif 'Yearly' in str(value):
                cell.fill = PatternFill(start_color='DEB887', end_color='DEB887', fill_type='solid')

for col_letter, width in column_widths.items():
    ws_recurring.column_dimensions[col_letter].width = width
ws_recurring.freeze_panes = 'A2'
ws_recurring.auto_filter.ref = f"A1:N{len(recurring) + 1}"

# --- PROFESSIONAL SHEET ---
ws_prof = wb.create_sheet("Professional")
professional = df[df['Category'] == 'Professional']

for col_num, header in enumerate(headers, 1):
    cell = ws_prof.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color='4A90A4', end_color='4A90A4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for row_num, row_data in enumerate(professional.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_prof.cell(row=row_num, column=col_num, value=value if pd.notna(value) else '')
        cell.font = body_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border

for col_letter, width in column_widths.items():
    ws_prof.column_dimensions[col_letter].width = width
ws_prof.freeze_panes = 'A2'
ws_prof.auto_filter.ref = f"A1:N{len(professional) + 1}"

# --- PERSONAL SHEET ---
ws_pers = wb.create_sheet("Personal")
personal = df[df['Category'] == 'Personal']

for col_num, header in enumerate(headers, 1):
    cell = ws_pers.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color='9B59B6', end_color='9B59B6', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for row_num, row_data in enumerate(personal.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_pers.cell(row=row_num, column=col_num, value=value if pd.notna(value) else '')
        cell.font = body_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border

for col_letter, width in column_widths.items():
    ws_pers.column_dimensions[col_letter].width = width
ws_pers.freeze_panes = 'A2'
ws_pers.auto_filter.ref = f"A1:N{len(personal) + 1}"

# --- CRITICAL & HIGH PRIORITY SHEET ---
ws_priority = wb.create_sheet("High Priority")
high_priority = df[df['Priority'].isin(['Critical', 'High'])]

for col_num, header in enumerate(headers, 1):
    cell = ws_priority.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for row_num, row_data in enumerate(high_priority.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_priority.cell(row=row_num, column=col_num, value=value if pd.notna(value) else '')
        cell.font = body_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border
        # Priority coloring
        if col_num == 8:
            if value == 'Critical':
                cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                cell.font = Font(name='Aptos', size=10, bold=True, color='FFFFFF')
            elif value == 'High':
                cell.fill = PatternFill(start_color='FFA94D', end_color='FFA94D', fill_type='solid')
                cell.font = Font(name='Aptos', size=10, bold=True)

for col_letter, width in column_widths.items():
    ws_priority.column_dimensions[col_letter].width = width
ws_priority.freeze_panes = 'A2'
ws_priority.auto_filter.ref = f"A1:N{len(high_priority) + 1}"

# --- CONTEXT VIEWS SHEET ---
ws_context = wb.create_sheet("By Context")

# Get unique contexts
contexts = df['Context'].dropna().unique()
current_row = 1

for context in sorted(contexts):
    # Context header
    cell = ws_context.cell(row=current_row, column=1, value=context)
    cell.font = Font(name='Aptos', size=12, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    ws_context.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    current_row += 1
    
    # Sub-headers
    sub_headers = ['Item', 'Project/Area', 'Priority', 'Status', 'Next Action']
    for col_num, header in enumerate(sub_headers, 1):
        cell = ws_context.cell(row=current_row, column=col_num, value=header)
        cell.font = Font(name='Aptos', size=10, bold=True)
        cell.fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
        cell.border = thin_border
    current_row += 1
    
    # Context items
    context_items = df[df['Context'] == context]
    for _, row in context_items.iterrows():
        ws_context.cell(row=current_row, column=1, value=row['Item']).font = body_font
        ws_context.cell(row=current_row, column=2, value=row['Project/Area']).font = body_font
        ws_context.cell(row=current_row, column=3, value=row['Priority']).font = body_font
        ws_context.cell(row=current_row, column=4, value=row['Status']).font = body_font
        ws_context.cell(row=current_row, column=5, value=row['Next Action']).font = body_font
        for col in range(1, 6):
            ws_context.cell(row=current_row, column=col).border = thin_border
        current_row += 1
    
    current_row += 1  # Space between contexts

ws_context.column_dimensions['A'].width = 55
ws_context.column_dimensions['B'].width = 30
ws_context.column_dimensions['C'].width = 10
ws_context.column_dimensions['D'].width = 18
ws_context.column_dimensions['E'].width = 50

# --- WEEKLY REVIEW SHEET ---
ws_review = wb.create_sheet("Weekly Review")

review_sections = [
    ("🔴 CRITICAL ITEMS - Address Immediately", df[df['Priority'] == 'Critical']),
    ("🟠 HIGH PRIORITY - This Week", df[df['Priority'] == 'High']),
    ("⏳ WAITING FOR - Follow Up", df[df['Status'] == 'Waiting For']),
    ("📋 ACTIVE PROJECTS - Review Progress", df[df['Status'] == 'Project']),
    ("🔄 RECURRING - Daily Tasks", df[df['Status'] == 'Recurring - Daily']),
    ("🔄 RECURRING - Weekly Tasks", df[df['Status'] == 'Recurring - Weekly']),
]

current_row = 1

# Title
cell = ws_review.cell(row=current_row, column=1, value="WEEKLY REVIEW - Shreyas GTD")
cell.font = Font(name='Aptos', size=14, bold=True)
current_row += 2

for section_title, section_data in review_sections:
    if len(section_data) == 0:
        continue
        
    # Section header
    cell = ws_review.cell(row=current_row, column=1, value=section_title)
    cell.font = Font(name='Aptos', size=11, bold=True)
    ws_review.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 1
    
    # Column headers
    cols = ['Item', 'Project/Area', 'Next Action', 'Status']
    for col_num, header in enumerate(cols, 1):
        cell = ws_review.cell(row=current_row, column=col_num, value=header)
        cell.font = Font(name='Aptos', size=10, bold=True)
        cell.fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
        cell.border = thin_border
    current_row += 1
    
    # Data
    for _, row in section_data.iterrows():
        ws_review.cell(row=current_row, column=1, value=row['Item']).font = body_font
        ws_review.cell(row=current_row, column=2, value=row['Project/Area']).font = body_font
        ws_review.cell(row=current_row, column=3, value=row['Next Action']).font = body_font
        ws_review.cell(row=current_row, column=4, value=row['Status']).font = body_font
        for col in range(1, 5):
            ws_review.cell(row=current_row, column=col).border = thin_border
        current_row += 1
    
    current_row += 1

ws_review.column_dimensions['A'].width = 55
ws_review.column_dimensions['B'].width = 30
ws_review.column_dimensions['C'].width = 50
ws_review.column_dimensions['D'].width = 18

# --- STRATEGIC ACCOUNTS SHEET ---
ws_accounts = wb.create_sheet("Strategic Accounts")

accounts = ['World Bank Group', 'HPE', 'CAA', 'BAT', 'Riyadh Air', 'AstraZeneca']
current_row = 1

# Title
cell = ws_accounts.cell(row=current_row, column=1, value="STRATEGIC ACCOUNT PORTFOLIO")
cell.font = Font(name='Aptos', size=14, bold=True)
current_row += 2

for account in accounts:
    account_items = df[df['Project/Area'].str.contains(account, case=False, na=False)]
    
    if len(account_items) == 0:
        continue
    
    # Account header
    cell = ws_accounts.cell(row=current_row, column=1, value=f"📊 {account}")
    cell.font = Font(name='Aptos', size=12, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4A90A4', end_color='4A90A4', fill_type='solid')
    ws_accounts.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 1
    
    # Column headers
    cols = ['Item', 'Status', 'Priority', 'Notes']
    for col_num, header in enumerate(cols, 1):
        cell = ws_accounts.cell(row=current_row, column=col_num, value=header)
        cell.font = Font(name='Aptos', size=10, bold=True)
        cell.fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
        cell.border = thin_border
    current_row += 1
    
    # Items
    for _, row in account_items.iterrows():
        ws_accounts.cell(row=current_row, column=1, value=row['Item']).font = body_font
        ws_accounts.cell(row=current_row, column=2, value=row['Status']).font = body_font
        ws_accounts.cell(row=current_row, column=3, value=row['Priority']).font = body_font
        ws_accounts.cell(row=current_row, column=4, value=row['Notes']).font = body_font
        for col in range(1, 5):
            ws_accounts.cell(row=current_row, column=col).border = thin_border
        current_row += 1
    
    current_row += 1

ws_accounts.column_dimensions['A'].width = 55
ws_accounts.column_dimensions['B'].width = 18
ws_accounts.column_dimensions['C'].width = 10
ws_accounts.column_dimensions['D'].width = 50

# --- ROLE DEVELOPMENT SHEET ---
ws_role = wb.create_sheet("Role Development")

role_areas = [
    'AI Integration',
    'Strategic Accounts',
    'Innovation Lab',
    'Presales-Delivery Bridge'
]
current_row = 1

# Title
cell = ws_role.cell(row=current_row, column=1, value="ROLE DEVELOPMENT TASKS - Strategic Innovation Lead")
cell.font = Font(name='Aptos', size=14, bold=True)
current_row += 2

for area in role_areas:
    area_items = df[df['Project/Area'].str.contains(area, case=False, na=False)]
    
    if len(area_items) == 0:
        continue
    
    # Area header
    cell = ws_role.cell(row=current_row, column=1, value=f"🎯 {area}")
    cell.font = Font(name='Aptos', size=12, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    ws_role.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 1
    
    # Column headers
    cols = ['Task', 'Frequency', 'Time Required', 'Notes']
    for col_num, header in enumerate(cols, 1):
        cell = ws_role.cell(row=current_row, column=col_num, value=header)
        cell.font = Font(name='Aptos', size=10, bold=True)
        cell.fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
        cell.border = thin_border
    current_row += 1
    
    # Items
    for _, row in area_items.iterrows():
        ws_role.cell(row=current_row, column=1, value=row['Item']).font = body_font
        ws_role.cell(row=current_row, column=2, value=row['Status']).font = body_font
        ws_role.cell(row=current_row, column=3, value=row['Time Required']).font = body_font
        ws_role.cell(row=current_row, column=4, value=row['Notes']).font = body_font
        for col in range(1, 5):
            ws_role.cell(row=current_row, column=col).border = thin_border
        current_row += 1
    
    current_row += 1

ws_role.column_dimensions['A'].width = 55
ws_role.column_dimensions['B'].width = 20
ws_role.column_dimensions['C'].width = 15
ws_role.column_dimensions['D'].width = 50

# --- FAMILY & LIFE SHEET ---
ws_family = wb.create_sheet("Family & Life")

family_areas = [
    ('👨‍👩‍👧 Family - Partner (MMD)', 'Family - Partner'),
    ('👶 Family - Children (Ananya)', 'Family - Children'),
    ('👴👵 Family - Parents', 'Family - Parents'),
    ('🏠 Home', 'Home'),
    ('❤️ Health', 'Health'),
    ('💰 Financial', 'Financial'),
    ('🎯 Goals & Planning', 'Goals'),
]

current_row = 1

# Title
cell = ws_family.cell(row=current_row, column=1, value="FAMILY & LIFE PRIORITIES")
cell.font = Font(name='Aptos', size=14, bold=True)
current_row += 2

for display_name, search_term in family_areas:
    area_items = df[df['Project/Area'].str.contains(search_term, case=False, na=False)]
    
    if len(area_items) == 0:
        continue
    
    # Area header
    cell = ws_family.cell(row=current_row, column=1, value=display_name)
    cell.font = Font(name='Aptos', size=12, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='9B59B6', end_color='9B59B6', fill_type='solid')
    ws_family.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 1
    
    # Column headers
    cols = ['Item', 'Priority', 'Status', 'Notes']
    for col_num, header in enumerate(cols, 1):
        cell = ws_family.cell(row=current_row, column=col_num, value=header)
        cell.font = Font(name='Aptos', size=10, bold=True)
        cell.fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
        cell.border = thin_border
    current_row += 1
    
    # Items
    for _, row in area_items.iterrows():
        ws_family.cell(row=current_row, column=1, value=row['Item']).font = body_font
        ws_family.cell(row=current_row, column=2, value=row['Priority']).font = body_font
        ws_family.cell(row=current_row, column=3, value=row['Status']).font = body_font
        ws_family.cell(row=current_row, column=4, value=row['Notes']).font = body_font
        for col in range(1, 5):
            ws_family.cell(row=current_row, column=col).border = thin_border
        current_row += 1
    
    current_row += 1

ws_family.column_dimensions['A'].width = 55
ws_family.column_dimensions['B'].width = 10
ws_family.column_dimensions['C'].width = 18
ws_family.column_dimensions['D'].width = 50

# --- BUSINESS PLANNING SHEET ---
ws_business = wb.create_sheet("Business Planning")

current_row = 1

# Title
cell = ws_business.cell(row=current_row, column=1, value="🚀 NEW BUSINESS PLANNING - Q1 2026 Goal")
cell.font = Font(name='Aptos', size=14, bold=True)
current_row += 2

# Get business planning items
business_items = df[df['Project/Area'].str.contains('Business Planning|Personal Projects|Professional Identity|Personal Branding', case=False, na=False, regex=True)]

# Column headers
cols = ['Item', 'Project Area', 'Priority', 'Status', 'Next Action', 'Notes']
for col_num, header in enumerate(cols, 1):
    cell = ws_business.cell(row=current_row, column=col_num, value=header)
    cell.font = Font(name='Aptos', size=10, bold=True)
    cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    cell.font = Font(name='Aptos', size=10, bold=True, color='FFFFFF')
    cell.border = thin_border
current_row += 1

for _, row in business_items.iterrows():
    ws_business.cell(row=current_row, column=1, value=row['Item']).font = body_font
    ws_business.cell(row=current_row, column=2, value=row['Project/Area']).font = body_font
    ws_business.cell(row=current_row, column=3, value=row['Priority']).font = body_font
    ws_business.cell(row=current_row, column=4, value=row['Status']).font = body_font
    ws_business.cell(row=current_row, column=5, value=row['Next Action']).font = body_font
    ws_business.cell(row=current_row, column=6, value=row['Notes']).font = body_font
    for col in range(1, 7):
        ws_business.cell(row=current_row, column=col).border = thin_border
    # Highlight critical items
    if row['Priority'] == 'Critical':
        for col in range(1, 7):
            ws_business.cell(row=current_row, column=col).fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
    current_row += 1

ws_business.column_dimensions['A'].width = 45
ws_business.column_dimensions['B'].width = 20
ws_business.column_dimensions['C'].width = 10
ws_business.column_dimensions['D'].width = 15
ws_business.column_dimensions['E'].width = 40
ws_business.column_dimensions['F'].width = 35

# --- LEGEND SHEET ---
ws_legend = wb.create_sheet("Legend & Guide")

current_row = 1

# Title
cell = ws_legend.cell(row=current_row, column=1, value="GTD LEGEND & COLOR GUIDE")
cell.font = Font(name='Aptos', size=14, bold=True)
current_row += 2

# Priority Legend
cell = ws_legend.cell(row=current_row, column=1, value="PRIORITY COLORS")
cell.font = Font(name='Aptos', size=12, bold=True)
current_row += 1

priority_legend = [
    ('Critical', 'FF6B6B', 'Urgent and important - address immediately'),
    ('High', 'FFA94D', 'Important - complete this week'),
    ('Medium', 'FFD93D', 'Significant - schedule appropriately'),
    ('Low', '6BCB77', 'Nice to have - do when time permits'),
]

for priority, color, desc in priority_legend:
    cell = ws_legend.cell(row=current_row, column=1, value=priority)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.font = body_font
    cell.border = thin_border
    ws_legend.cell(row=current_row, column=2, value=desc).font = body_font
    current_row += 1

current_row += 1

# Status Legend
cell = ws_legend.cell(row=current_row, column=1, value="STATUS COLORS")
cell.font = Font(name='Aptos', size=12, bold=True)
current_row += 1

status_legend = [
    ('Next Action', 'FF6B6B', 'Single next physical action to take'),
    ('Project', '4ECDC4', 'Multi-step outcome requiring planning'),
    ('Waiting For', 'A8E6CF', 'Delegated or waiting for external input'),
    ('Someday/Maybe', 'DDA0DD', 'Ideas to review later'),
    ('Recurring - Daily', 'FFB6C1', 'Daily habits and routines'),
    ('Recurring - Weekly', 'E6E6FA', 'Weekly recurring tasks'),
    ('Recurring - Monthly', 'B0E0E6', 'Monthly recurring tasks'),
    ('Recurring - Quarterly', 'F0E68C', 'Quarterly reviews and tasks'),
    ('Ongoing', '87CEEB', 'Continuous improvement areas'),
]

for status, color, desc in status_legend:
    cell = ws_legend.cell(row=current_row, column=1, value=status)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.font = body_font
    cell.border = thin_border
    ws_legend.cell(row=current_row, column=2, value=desc).font = body_font
    current_row += 1

current_row += 1

# Context Legend
cell = ws_legend.cell(row=current_row, column=1, value="CONTEXTS")
cell.font = Font(name='Aptos', size=12, bold=True)
current_row += 1

context_legend = [
    ('@Computer', 'Tasks requiring laptop/desktop'),
    ('@Phone', 'Calls and phone-based tasks'),
    ('@Office', 'Tasks requiring office presence'),
    ('@Home', 'Tasks to do at home'),
    ('@Errands', 'Out-and-about tasks'),
    ('@Thinking', 'Strategic thinking and planning'),
    ('@Reading', 'Reading and study time'),
    ('@Anywhere', 'Can be done anywhere'),
    ('@Gym', 'Exercise and fitness tasks'),
]

for context, desc in context_legend:
    ws_legend.cell(row=current_row, column=1, value=context).font = body_font
    ws_legend.cell(row=current_row, column=1).border = thin_border
    ws_legend.cell(row=current_row, column=2, value=desc).font = body_font
    current_row += 1

current_row += 2

# GTD Workflow
cell = ws_legend.cell(row=current_row, column=1, value="GTD WORKFLOW REMINDER")
cell.font = Font(name='Aptos', size=12, bold=True)
current_row += 1

gtd_steps = [
    "1. CAPTURE - Collect what has your attention (✓ Done with trigger list)",
    "2. CLARIFY - Process what each item means",
    "3. ORGANIZE - Put items where they belong",
    "4. REFLECT - Review frequently (Weekly Review)",
    "5. ENGAGE - Simply do (use Next Actions list)",
]

for step in gtd_steps:
    ws_legend.cell(row=current_row, column=1, value=step).font = body_font
    current_row += 1

current_row += 2

# Personal Note
cell = ws_legend.cell(row=current_row, column=1, value="PERSONAL CONTEXT NOTES")
cell.font = Font(name='Aptos', size=12, bold=True)
current_row += 1

personal_notes = [
    "• 9-month-old daughter (Ananya) - be present, limit device time",
    "• Wife (MMD) - prioritize relationship and support",
    "• Work boundaries - flexible but limited hours to avoid burnout",
    "• Business goal - complete plan Q1 2026, start after leaving job",
    "• Strategic Innovation Lead role at Upside Learning/UpsideLogic",
    "• Key accounts: World Bank, HPE, Riyadh Air, CAA, BAT, AstraZeneca",
]

for note in personal_notes:
    ws_legend.cell(row=current_row, column=1, value=note).font = body_font
    current_row += 1

ws_legend.column_dimensions['A'].width = 25
ws_legend.column_dimensions['B'].width = 50

# Save workbook
output_path = '/Users/shreyas/Downloads/GTD-Gather/Shreyas_GTD_Master_List.xlsx'
wb.save(output_path)

print(f"✅ GTD Excel file created successfully!")
print(f"📁 Location: {output_path}")
print(f"\n📊 Sheets created:")
print("   1. GTD Master List - Complete list with all items")
print("   2. Next Actions - Immediate action items")
print("   3. Projects - Multi-step outcomes")
print("   4. Waiting For - Delegated/pending items")
print("   5. Someday-Maybe - Future ideas")
print("   6. Recurring Tasks - Daily/Weekly/Monthly tasks")
print("   7. Professional - Work-related items")
print("   8. Personal - Personal life items")
print("   9. High Priority - Critical and High priority items")
print("  10. By Context - Items grouped by context (@Computer, @Phone, etc.)")
print("  11. Weekly Review - Formatted for weekly GTD review")
print("  12. Strategic Accounts - Account-specific view")
print("  13. Role Development - Tasks by role responsibility area")
print("  14. Family & Life - Personal priorities organized")
print("  15. Business Planning - New business venture planning")
print("  16. Legend & Guide - Color codes and GTD reference")
print(f"\n📝 Total items captured: {len(df)}")
print(f"   - Professional: {len(df[df['Category'] == 'Professional'])}")
print(f"   - Personal: {len(df[df['Category'] == 'Personal'])}")
print(f"\n🎯 Ready for GTD processing!")
